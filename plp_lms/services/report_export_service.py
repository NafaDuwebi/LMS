from io import BytesIO
from typing import List, Dict
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_report_pdf(title: str, headers: List[str], rows: List[Dict]) -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4) if len(headers) > 6 else A4)
    elements = []
    primary = HexColor("#1A2B4A")
    accent = HexColor("#C9912B")

    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], textColor=primary, fontSize=16, spaceAfter=12)
    elements.append(Paragraph(title, title_style))

    data = [headers]
    for row in rows:
        data.append([str(row.get(h, "")) for h in headers])

    col_count = len(headers)
    col_widths = [doc.width / col_count] * col_count
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F9F9F9")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def generate_report_xlsx(title: str, headers: List[str], rows: List[Dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A2B4A", end_color="1A2B4A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            value = row.get(h, "")
            if isinstance(value, float):
                value = round(value, 1)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 2)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
